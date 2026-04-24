import openpyxl
import os
import json

# 读取Excel功能清单
wb = openpyxl.load_workbook(r'd:\CODE\project\community\惠民社区功能清单.xlsx')
ws = wb.active

# 提取所有功能点
features = []
current_core = None
current_sub = None
current_feature = None

for i in range(1, ws.max_row + 1):
    row_data = [cell.value for cell in ws[i]]
    core = row_data[1]
    sub = row_data[2]
    feature = row_data[3]
    sub_feature = row_data[4]
    desc = row_data[5]
    
    if core:
        current_core = core
    if sub:
        current_sub = sub
    if feature:
        current_feature = feature
        
    if desc and any(v is not None for v in [core, sub, feature, sub_feature]):
        features.append({
            'row': i,
            'core_module': current_core,
            'sub_module': current_sub,
            'feature': current_feature,
            'sub_feature': sub_feature,
            'description': desc
        })

# 前端已实现的页面
pages_dir = r'd:\CODE\project\community\pages'
implemented_pages = []
if os.path.exists(pages_dir):
    for d in os.listdir(pages_dir):
        page_path = os.path.join(pages_dir, d)
        if os.path.isdir(page_path):
            implemented_pages.append(d)

# 分包页面
subpackages = ['package-market', 'package-worker', 'package-merchant']
for pkg in subpackages:
    pkg_path = os.path.join(r'd:\CODE\project\community', pkg, 'pages')
    if os.path.exists(pkg_path):
        for d in os.listdir(pkg_path):
            page_path = os.path.join(pkg_path, d)
            if os.path.isdir(page_path):
                implemented_pages.append(f'{pkg}/{d}')

# 功能关键词映射
feature_keywords = {
    '注册': ['register', '注册'],
    '登录': ['login', '登录'],
    '忘记密码': ['forget-password', '忘记密码'],
    '首页': ['index', '首页'],
    '定位': ['location', '定位'],
    '搜索': ['search', '搜索'],
    '服务分类': ['classify', '分类'],
    '直约服务商': ['service-provider', '服务商'],
    '直约技工': ['worker', '技工'],
    '秒杀': ['miaosha', '秒杀'],
    '领券': ['coupon', '领券', '优惠券'],
    '家事币商城': ['benefit', '家事币', '惠民卡'],
    '家超市': ['push', '家超市', '本地商城'],
    '拼团': ['group', '拼团', 'combo'],
    '邻里帮帮': ['neighbor', '邻里'],
    '代取': ['take', '代取'],
    '接送小孩': ['child', '接送'],
    '陪诊': ['escort', '陪诊'],
    '陪读': ['study', '陪读'],
    '便民服务': ['recomm', '便民'],
    '邻里互动': ['community', '互动'],
    '二手闲置': ['secondhand', '二手', '闲置'],
    '消息': ['message', '消息'],
    '我的': ['user', '我的'],
    '个人资料': ['user-edit', '资料'],
    '我的收益': ['wallet', '收益', '钱包'],
    '我的订单': ['order', '订单'],
    '家集市订单': ['market-order', '家集市'],
    '推客订单': ['promoter', '推客'],
    '我的社区': ['my-posts', 'my-activities', '我的社区'],
    '地址管理': ['address', '地址'],
    '设置': ['settings', '设置'],
    '关于我们': ['about', '关于'],
    '反馈': ['feedback', '反馈'],
    '本地集市': ['market', '集市'],
    '商品详情': ['goods-detail', '商品'],
    '购物车': ['cart', '购物车'],
    '订单确认': ['order-confrim', '确认'],
    '活动': ['activity', '活动'],
    '聊天': ['chat', '聊天'],
    '入驻': ['join', '入驻'],
}

# 检查每个功能是否实现
implemented_features = []
not_implemented_features = []

for feat in features:
    desc = feat['description'] or ''
    sub_feature = feat['sub_feature'] or ''
    feature = feat['feature'] or ''
    
    is_implemented = False
    matched_page = None
    
    for keyword, page_names in feature_keywords.items():
        if keyword in desc or keyword in sub_feature or keyword in feature:
            for page in page_names:
                if any(page in p for p in implemented_pages):
                    is_implemented = True
                    matched_page = page
                    break
        if is_implemented:
            break
    
    if is_implemented:
        implemented_features.append({**feat, 'matched_page': matched_page})
    else:
        not_implemented_features.append(feat)

# 输出报告
print("=" * 80)
print("惠民社区功能清单 vs 前端实现对比报告")
print("=" * 80)
print(f"\n功能清单总条目数: {len(features)}")
print(f"已实现功能数: {len(implemented_features)}")
print(f"未实现/待确认功能数: {len(not_implemented_features)}")
print(f"前端页面数: {len(implemented_pages)}")

print("\n" + "=" * 80)
print("已实现功能列表:")
print("=" * 80)
for i, feat in enumerate(implemented_features, 1):
    print(f"{i}. [{feat['core_module'] or ''}] {feat['sub_feature'] or feat['feature'] or ''}")
    print(f"   描述: {feat['description'][:50]}...")
    print(f"   对应页面: {feat['matched_page']}")

print("\n" + "=" * 80)
print("未实现/待确认功能列表:")
print("=" * 80)
for i, feat in enumerate(not_implemented_features, 1):
    print(f"{i}. [{feat['core_module'] or ''}] {feat['sub_feature'] or feat['feature'] or ''}")
    print(f"   描述: {feat['description'][:60]}...")

# 保存详细报告
report = {
    'summary': {
        'total_features': len(features),
        'implemented': len(implemented_features),
        'not_implemented': len(not_implemented_features),
        'implemented_pages': implemented_pages
    },
    'implemented_features': implemented_features,
    'not_implemented_features': not_implemented_features
}

with open(r'd:\CODE\project\community\feature_comparison_report.json', 'w', encoding='utf-8') as f:
    json.dump(report, f, ensure_ascii=False, indent=2)

print("\n详细报告已保存到: d:\\CODE\\project\\community\\feature_comparison_report.json")
