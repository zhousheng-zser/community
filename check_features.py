import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

# ==================== 1. 读取功能清单 ====================
wb = openpyxl.load_workbook('惠民社区功能清单.xlsx')
ws = wb.active

rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    if any(c is not None and str(c).strip() for c in row[:6]):
        rows.append(row)

# 提取用户端小程序的所有功能模块
in_user_app = False
features = []
for row in rows:
    a = str(row[0]).strip() if row[0] else ''
    if '用户端小程序' in a:
        in_user_app = True
        continue
    if in_user_app and a and any(x in a for x in ['管理后台', '商家PC']):
        break
    if not in_user_app:
        continue
    b = str(row[1]).strip() if row[1] else ''
    c = str(row[2]).strip() if row[2] else ''
    d = str(row[3]).strip() if row[3] else ''
    e = str(row[4]).strip() if row[4] else ''
    parts = [p for p in [b, c, d, e] if p]
    if parts:
        features.append(' > '.join(parts))

# 提取惠民福利卡
in_welfare = False
for row in rows:
    a = str(row[0]).strip() if row[0] else ''
    if '惠民福利卡' in a:
        in_welfare = True
    if in_welfare and a and any(x in a for x in ['管理后台', '商家PC', '核心模块']):
        break
    if not in_welfare:
        continue
    b = str(row[1]).strip() if row[1] else ''
    c = str(row[2]).strip() if row[2] else ''
    d = str(row[3]).strip() if row[3] else ''
    e = str(row[4]).strip() if row[4] else ''
    parts = [p for p in [a, b, c, d, e] if p]
    if parts:
        features.append(' > '.join(parts))

# ==================== 2. 项目中的页面分类 ====================
user_pages = {
    'index', 'community', 'message', 'chat', 'community-publish', 'classify',
    'goods-cart', 'user', 'gorder-list', 'gorder-detail', 'cash', 'coupons-all',
    'coupons-my', 'book', 'audit', 'dispatch', 'user-edit', 'order-detail',
    'order-confrim', 'service-provider-shop', 'service', 'push-channel',
    'push-goods-list', 'push-product-detail', 'market-shop', 'market-banner',
    'worker-detail', 'merchant-detail', 'account', 'order', 'recomm', 'activity',
    'goods-classify', 'tidy-service', 'goods-detail', 'activity-list', 'user-code',
    'book-my', 'service-orders-my', 'service-order-detail',
    'neighbor-assist-orders-my', 'neighbor-assist-order-detail', 'goods-confrim',
    'book-detail', 'my-posts', 'push-video-list', 'push-video-category',
    'push-daily-news', 'push-top-sales', 'push-hot-video-list',
    'push-periodic-list', 'settings', 'feedback', 'my-follows', 'my-activities',
    'order-service', 'order-publish', 'join-worker', 'join-service', 'join-market',
    'address', 'about', 'market-order-detail', 'market-order-list', 'rider-location',
    'privacy-publish', 'privacy-reply', 'notification-settings', 'benefit-orders',
    'promoter-orders', 'combo-orders', 'appeal-list', 'activity-manage',
    'community-steward', 'platform-kefu', 'wallet-transactions', 'shopping-search',
    'consumer-rights', 'after-sale-apply', 'login', 'register', 'forget-password',
    'env-switcher', 'notification-center',
    'commission-overview', 'commission-records', 'partner-manage',
    'benefit-coin-balance', 'benefit-coin-detail', 'benefit-coin-mall',
    'group-detail', 'group-list',
}

worker_pages = {'worker-home', 'worker-mine', 'worker-order-detail', 'worker-orders'}
merchant_pages = {'merchant-home', 'merchant-service', 'merchant-goods', 'merchant-goods-edit',
    'merchant-orders', 'merchant-mine', 'merchant-ship', 'merchant-refund-handle',
    'merchant-dispatch', 'merchant-arbitrate'}
market_pages = {'market-home'}
sp_pages = {'sp-home', 'sp-orders', 'sp-dispatch', 'sp-services', 'sp-mine'}
rider_pages = {'rider-tasks', 'rider-home', 'rider-mine'}

all_pages = user_pages | worker_pages | merchant_pages | market_pages | sp_pages | rider_pages

# ==================== 3. 精确映射 ====================
mapping = {
    # 注册登录
    '注册 > 注册': ['register'],
    '注册 > 用户注册协议': ['register'],
    '登录 > 密码登录': ['login'],
    '登录 > 快捷登录': ['login'],
    '登录 > 随便逛逛': ['index'],
    '登录 > 忘记密码': ['forget-password'],

    # 首页
    '小程序首页 > 定位': ['index'],
    '小程序首页 > 搜索': ['index', 'shopping-search'],
    '小程序首页 > 宫格分类 > 便民服务 > 家修急事': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 家庭服务': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 闲置二手': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 家庭陪护': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 上门美业': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 宝宝家事': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 家庭装修': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 便民服务': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 家庭保洁': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 家居养护': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 保姆月嫂': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 上门回收': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 家电维修': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 家电清洗': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 衣物洗护': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 上门安装': ['service'],
    '小程序首页 > 宫格分类 > 便民服务 > 其它服务分类': ['service'],
    '小程序首页 > 其他服务 > 直约服务商': ['service-provider-shop'],
    '小程序首页 > 其他服务 > 直约技工': ['worker-detail'],
    '小程序首页 > 秒杀': ['push-channel'],
    '小程序首页 > 领券': ['coupons-all'],
    '小程序首页 > 家事币商城': ['benefit-coin-mall'],
    '小程序首页 > 家超市': ['index'],
    '小程序首页 > 特价拼团': ['group-list'],
    '小程序首页 > 易达速递': ['index'],
    '小程序首页 > 啄木鸟': ['index'],
    '小程序首页 > 榕益收': ['index'],

    # 团购
    '团购 > 团购列表': ['group-list'],
    '团购 > 团购详情': ['group-detail'],

    # 邻里帮帮
    '邻里帮帮 > 一键发布': ['order-publish'],
    '邻里帮帮 > 代取': ['order-publish'],
    '邻里帮帮 > 接送小孩': ['order-publish'],
    '邻里帮帮 > 陪诊': ['order-publish'],
    '邻里帮帮 > 陪读': ['order-publish'],

    # 便民服务信息板块
    '便民服务 > 轮播图': ['market-banner'],
    '便民服务 > 搜索': ['service'],
    '便民服务 > 分类 > 服务分类': ['classify', 'service'],
    '便民服务 > 分类 > 服务类 > 热门服务': ['service'],
    '便民服务 > 分类 > 服务类 > 邻里互助': ['service'],
    '便民服务 > 分类 > 服务类 > 维修服务': ['service'],
    '便民服务 > 分类 > 服务类 > 家政保洁': ['service'],
    '便民服务 > 分类 > 服务类 > 其它服务等': ['service'],
    '便民服务 > 分类 > 服务类 > 服务列表': ['service'],
    '便民服务 > 分类 > 服务类 > 服务详情': ['service-order-detail'],
    '便民服务 > 分类 > 信息类 > 热门信息': ['community'],
    '便民服务 > 分类 > 信息类 > 二手闲置': ['community'],
    '便民服务 > 分类 > 信息类 > 组织活动': ['activity-list'],
    '便民服务 > 分类 > 信息类 > 育儿心得': ['community'],
    '便民服务 > 分类 > 信息类 > 其它信息等': ['community'],
    '便民服务 > 分类 > 信息类 > 信息列表': ['community'],
    '便民服务 > 分类 > 信息类 > 信息详情': ['order-detail'],
    '便民服务 > 分类 > 最新发布': ['community'],
    '便民服务 > 发布服务 > 发布邻里互助': ['order-publish'],
    '便民服务 > 发布服务 > 发布家政服务': ['order-publish'],
    '便民服务 > 发布信息 > 发布闲置家具': ['order-publish'],
    '便民服务 > 发布信息 > 发布组织活动': ['community-publish'],
    '便民服务 > 发布信息 > 发布社区物业': ['community-publish'],

    # 商城
    '商城版块 > 搜索 > 搜索': ['shopping-search'],
    '商城版块 > 搜索 > 搜索历史': ['shopping-search'],
    '商城版块 > 搜索 > 清空': ['shopping-search'],
    '商城版块 > 搜索 > 猜你喜欢': ['shopping-search'],
    '商城版块 > 搜索结果 > 商品搜索': ['shopping-search'],
    '商城版块 > 搜索结果 > 店铺搜索': ['market-shop'],
    '商城版块 > 多宫格 > 食品生鲜': ['goods-classify'],
    '商城版块 > 多宫格 > 美妆洗护': ['goods-classify'],
    '商城版块 > 多宫格 > 居家百货': ['goods-classify'],
    '商城版块 > 多宫格 > 服装箱包': ['goods-classify'],
    '商城版块 > 多宫格 > 母婴系列': ['goods-classify'],
    '商城版块 > 多宫格 > 家用电器': ['goods-classify'],
    '商城版块 > 多宫格 > 数码产品': ['goods-classify'],
    '商城版块 > 多宫格 > 珠宝饰品': ['goods-classify'],
    '商城版块 > 多宫格 > 旅游出行': ['goods-classify'],
    '商城版块 > 多宫格 > 传统工艺': ['goods-classify'],
    '商城版块 > 营销活动 > 营销 > 促销信息': ['push-channel'],
    '商城版块 > 营销活动 > 营销 > 团购': ['group-list'],
    '商城版块 > 营销活动 > 营销 > 优惠券': ['coupons-all'],
    '商城版块 > 分类营销 > 公益扶贫': ['push-channel'],
    '商城版块 > 分类营销 > 打折促销': ['push-channel'],
    '商城版块 > 分类营销 > 家乡味道': ['push-channel'],
    '商城版块 > 分类营销 > 今日推荐': ['push-channel'],
    '商城版块 > 分类营销 > 爆款好货': ['push-channel'],
    '商城版块 > 分类营销 > 新品上架': ['push-channel'],
    '商城版块 > 推荐商品 > 分类': ['recomm'],
    '商城版块 > 推荐商品 > 筛选': ['recomm'],
    '商城版块 > 推荐商品 > 商品列表': ['recomm'],
    '商城版块 > 搜索 > 搜索': ['shopping-search'],
    '商城版块 > 商城分类 > 父分类': ['goods-classify'],
    '商城版块 > 商城分类 > 子分类': ['goods-classify'],
    '商城版块 > 购物车 > 管理': ['goods-cart'],
    '商城版块 > 购物车 > 删除': ['goods-cart'],
    '商城版块 > 为你推荐': ['recomm'],

    # 商品详情
    '商城商品详情 > 商品详情': ['goods-detail'],
    '商城商品详情 > 分享': ['goods-detail'],
    '商城商品详情 > 收藏': ['goods-detail'],
    '商城商品详情 > 购物车': ['goods-cart'],
    '商城商品详情 > 选择规格': ['goods-detail'],
    '商城商品详情 > 参数': ['goods-detail'],
    '商城商品详情 > 评论': ['goods-detail'],
    '商城商品详情 > 加入购物车': ['goods-detail'],
    '商城商品详情 > 立即购买': ['goods-detail'],

    # 确认订单
    '确认订单 > 收货地址': ['address'],
    '确认订单 > 商品信息': ['order-confrim', 'goods-confrim'],
    '确认订单 > 优惠券': ['order-confrim', 'goods-confrim'],
    '确认订单 > 上门自提订单': ['order-confrim'],
    '确认订单 > 支付': ['order-confrim'],
    '确认订单 > 付款成功': ['order-confrim'],

    # 聊天
    '聊天版块 > IM聊天 > 私聊 > 聊天': ['chat'],
    '聊天版块 > IM聊天 > 私聊 > 关注': ['chat'],
    '聊天版块 > IM聊天 > 私聊 > 取关': ['chat'],
    '聊天版块 > IM聊天 > 私聊 > 举报': ['feedback'],
    '聊天版块 > IM聊天 > 群聊 > 入群': ['chat'],
    '聊天版块 > IM聊天 > 群聊 > 群聊详情': ['chat'],
    '聊天版块 > IM聊天 > 群聊 > 退出群聊': ['chat'],
    '聊天版块 > IM聊天 > 群聊 > 举报': ['feedback'],

    # 订单
    '订单版块 > 商城订单 > 待付款': ['gorder-list'],
    '订单版块 > 商城订单 > 待付款订单详情': ['gorder-detail'],
    '订单版块 > 商城订单 > 待发货': ['gorder-list'],
    '订单版块 > 商城订单 > 待发货订单详情': ['gorder-detail'],
    '订单版块 > 商城订单 > 待收货': ['gorder-list'],
    '订单版块 > 商城订单 > 待收货订单详情': ['gorder-detail'],
    '订单版块 > 商城订单 > 待评价': ['gorder-list'],
    '订单版块 > 商城订单 > 待评价订单详情': ['gorder-detail'],
    '订单版块 > 商城订单 > 售后': ['gorder-list'],
    '订单版块 > 商城订单 > 售后订单详情 > 待退款详情': ['gorder-detail'],
    '订单版块 > 商城订单 > 售后订单详情 > 拒绝退款详情': ['gorder-detail'],
    '订单版块 > 商城订单 > 售后订单详情 > 退款成功详情': ['gorder-detail'],
    '订单版块 > 便民服务订单 > 全部': ['service-orders-my'],
    '订单版块 > 便民服务订单 > 待接单': ['service-orders-my'],
    '订单版块 > 便民服务订单 > 进行中': ['service-orders-my'],
    '订单版块 > 便民服务订单 > 已完成': ['service-orders-my'],
    '订单版块 > 便民服务订单 > 服务订单详情': ['service-order-detail'],

    # 社区
    '社区 > 定位 > LSB定位': ['community'],
    '社区 > 搜索 > 搜索输入': ['community'],
    '社区 > 公告 > 全部公告': ['notification-center'],
    '社区 > 公告 > 一键已读': ['notification-center'],
    '社区 > 发布 > 发布动态': ['community-publish'],
    '社区 > 发布 > 组织活动': ['community-publish'],
    '社区 > 热门话题 > 话题列表': ['push-channel'],
    '社区 > 热门活动 > 活动列表': ['activity-list'],
    '社区 > 邻里互动 > 互动列表': ['community'],

    # 一键发布
    '一键发布 > 轮播图 > 轮播图展示': ['market-banner'],
    '一键发布 > 代取 > 填写代取信息': ['order-publish'],
    '一键发布 > 接送小孩 > 填写接送小孩信息': ['order-publish'],
    '一键发布 > 陪诊 > 填写陪诊信息': ['order-publish'],
    '一键发布 > 陪读 > 填写陪读信息': ['order-publish'],

    # 消息
    '消息 > 消息通知 > 订单消息': ['message'],
    '消息 > 消息通知 > 系统消息': ['message'],

    # 我的
    '我的 > 个人资料 > 个人信息展示': ['user'],
    '我的 > 个人资料 > 修改': ['user-edit'],
    '我的 > 优惠券': ['coupons-my'],
    '我的 > 我的收益 > 金额 > 展示金额': ['cash'],
    '我的 > 我的收益 > 金额 > 可提现金额': ['cash'],
    '我的 > 我的收益 > 金额 > 提现方式': ['cash'],
    '我的 > 我的收益 > 金额 > 提现金额': ['cash'],
    '我的 > 我的收益 > 交易明细': ['wallet-transactions'],
    '我的 > 我的收益 > 提现明细': ['wallet-transactions'],
    '我的 > 我的订单 > 服务产品订单': ['service-orders-my'],
    '我的 > 我的订单 > 一键发布订单': ['neighbor-assist-orders-my'],
    '我的 > 我的订单 > 家集市订单': ['market-order-list'],
    '我的 > 我的订单 > 推客订单': ['promoter-orders'],
    '我的 > 我的社区 > 我的帖子': ['my-posts'],
    '我的 > 我的社区 > 我的关注 > 用户': ['my-follows'],
    '我的 > 我的社区 > 我的关注 > 帖子': ['my-follows'],
    '我的 > 我的社区 > 我的点赞 > 活动': ['my-activities'],
    '我的 > 我的社区 > 我的点赞 > 话题': ['my-activities'],
    '我的 > 我的社区 > 我的点赞 > 帖子': ['my-activities'],
    '我的 > 我的社区 > 参与话题': ['my-activities'],
    '我的 > 我的社区 > 参与活动': ['my-activities'],
    '我的 > 我的社区 > 诉求列表 > 全部': ['appeal-list'],
    '我的 > 我的社区 > 诉求列表 > 受理中': ['appeal-list'],
    '我的 > 我的社区 > 诉求列表 > 已结案': ['appeal-list'],
    '我的 > 加入惠民社区 > 技工入驻': ['join-worker'],
    '我的 > 加入惠民社区 > 家集市商家入驻': ['join-market'],
    '我的 > 加入惠民社区 > 服务商入驻': ['join-service'],
    '我的 > 其他服务 > 帮助反馈': ['feedback'],
    '我的 > 其他服务 > 在线小区管家': ['platform-kefu'],
    '我的 > 其他服务 > 关于我们': ['about'],
    '我的 > 地址管理 > 收货地址列表': ['address'],
    '我的 > 地址管理 > 新增收货地址': ['address'],
    '我的 > 平台客服': ['platform-kefu'],
    '我的 > 设置 > 修改资料': ['settings'],
    '我的 > 设置 > 修改密码': ['settings'],
    '我的 > 设置 > 换绑手机': ['settings'],
    '我的 > 设置 > 谁可以看我的发布': ['settings'],
    '我的 > 设置 > 谁可以看我的回复': ['settings'],
    '我的 > 设置 > 检查更新': ['settings'],
    '我的 > 设置 > 注销账号': ['settings'],
    '我的 > 设置 > 退出登录': ['settings'],

    # 惠民福利卡
    '惠民福利卡 > 购物省钱 > 链接多平台推送性价比好物': ['benefit-coin-mall'],
    '惠民福利卡 > 赚取佣金': ['commission-overview'],
    '惠民福利卡 > 惠民币商城 > 消费可得币，兑换奖励或抵扣': ['benefit-coin-mall'],
}

# 精确匹配
found_pages = set()
missing_features = []
matched_features = []

for feat in features:
    # 先尝试精确匹配
    if feat in mapping:
        pages = mapping[feat]
        found_pages.update(pages)
        matched_features.append(feat)
        continue

    # 模糊匹配
    matched = False
    for key, pages in mapping.items():
        if key in feat or feat in key:
            found_pages.update(pages)
            matched_features.append(feat)
            matched = True
            break

    if not matched:
        # 尝试根据关键词匹配
        for page in user_pages:
            page_name = page.replace('-', '')
            if page_name in feat.replace(' ', '').replace('>', '') or \
               any(part in feat for part in page.split('-')):
                found_pages.add(page)
                matched_features.append(feat)
                matched = True
                break

    if not matched:
        missing_features.append(feat)

print('=== 惠民社区前端功能检查报告 ===\n')
print('【功能清单统计】')
print(f'  功能清单总条目数: {len(features)}')
print(f'  已匹配功能数: {len(matched_features)}')
print(f'  未匹配功能数: {len(missing_features)}')
print()
print('【项目页面统计】')
print(f'  用户端总页面数: {len(user_pages)}')
print(f'  技工端页面数: {len(worker_pages)}')
print(f'  商家端页面数: {len(merchant_pages)}')
print(f'  集市端页面数: {len(market_pages)}')
print(f'  服务商端页面数: {len(sp_pages)}')
print(f'  骑手端页面数: {len(rider_pages)}')
print(f'  已覆盖用户端页面数: {len(found_pages & user_pages)} / {len(user_pages)}')
print()

if missing_features:
    print('--- 未匹配功能（建议检查是否缺失）---')
    for f in missing_features:
        print(f'  ? {f}')
    print()

uncovered_user = user_pages - found_pages
if uncovered_user:
    print(f'--- 未与功能清单关联的用户端页面({len(uncovered_user)}个) ---')
    for p in sorted(uncovered_user):
        print(f'  - {p}')
    print()

# 按模块分类报告
print('=== 按模块分类检查 ===\n')
modules = {
    '注册登录': ['register', 'login', 'forget-password'],
    '小程序首页': ['index', 'shopping-search', 'service', 'service-provider-shop', 'worker-detail',
                 'push-channel', 'coupons-all', 'benefit-coin-mall', 'group-list'],
    '团购': ['group-list', 'group-detail'],
    '邻里帮帮': ['order-publish'],
    '便民服务': ['market-banner', 'service', 'classify', 'community', 'order-publish',
                'community-publish', 'activity-list', 'service-order-detail', 'order-detail'],
    '商城版块': ['shopping-search', 'goods-classify', 'push-channel', 'recomm', 'goods-cart',
                'goods-detail', 'order-confrim', 'goods-confrim', 'address', 'market-shop'],
    '聊天版块': ['chat', 'feedback'],
    '订单版块': ['gorder-list', 'gorder-detail', 'service-orders-my', 'service-order-detail'],
    '社区': ['community', 'notification-center', 'community-publish', 'push-channel', 'activity-list'],
    '一键发布': ['market-banner', 'order-publish'],
    '消息': ['message'],
    '我的': ['user', 'user-edit', 'coupons-my', 'cash', 'wallet-transactions',
            'service-orders-my', 'neighbor-assist-orders-my', 'market-order-list',
            'promoter-orders', 'my-posts', 'my-follows', 'my-activities',
            'appeal-list', 'join-worker', 'join-market', 'join-service',
            'feedback', 'platform-kefu', 'about', 'address', 'settings'],
    '惠民福利卡': ['benefit-coin-mall', 'commission-overview', 'benefit-coin-balance',
                 'benefit-coin-detail'],
}

for mod_name, mod_pages in modules.items():
    existing = [p for p in mod_pages if p in user_pages]
    missing_in_mod = [p for p in mod_pages if p not in user_pages]
    status = 'OK' if not missing_in_mod else '部分缺失'
    print(f'[{mod_name}] {status} - 已有{len(existing)}个页面')
    if missing_in_mod:
        for p in missing_in_mod:
            print(f'  ! 缺失: {p}')
