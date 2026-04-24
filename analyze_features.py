import openpyxl
import json

wb = openpyxl.load_workbook(r'd:\CODE\project\community\惠民社区功能清单.xlsx')
ws = wb.active

# 提取所有功能点
features = []
for i in range(1, ws.max_row + 1):
    row_data = [cell.value for cell in ws[i]]
    # 提取有实际内容的行
    if any(v is not None for v in row_data[1:]):
        features.append({
            'row': i,
            'core_module': row_data[1],
            'sub_module': row_data[2],
            'feature': row_data[3],
            'sub_feature': row_data[4],
            'description': row_data[5]
        })

# 输出为JSON方便分析
print(json.dumps(features, ensure_ascii=False, indent=2))
